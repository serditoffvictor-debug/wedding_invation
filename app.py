from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, timezone, timedelta
import sqlite3
import io
import os


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data" / "guests.db"

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-key")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS guests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            attendance TEXT NOT NULL,
            companion TEXT NOT NULL,
            alcohol TEXT,
            transfer TEXT,
            allergies TEXT,
            children TEXT,
            comments TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.commit()
    conn.close()


# Важно: база создаётся не только при python app.py,
# но и при запуске через gunicorn / другой WSGI-сервер.
init_db()


@app.get("/")
def index():
    return render_template("index.html")


@app.get("/rsvp/health")
def rsvp_health():
    try:
        conn = get_db()
        conn.execute("SELECT 1").fetchone()
        conn.close()

        return jsonify(
            {
                "ok": True,
                "message": "RSVP backend работает",
                "database": str(DB_PATH),
            }
        )

    except Exception:
        app.logger.exception("Ошибка проверки RSVP")

        return jsonify(
            {
                "ok": False,
                "message": "RSVP backend не может открыть базу данных",
            }
        ), 500


@app.post("/rsvp")
def rsvp():
    name = request.form.get("name", "").strip()
    attendance = request.form.get("attendance", "").strip()

    # Поддерживаем и текущее имя поля guest_count,
    # и старое companion на случай, если где-то останется старый HTML.
    companion = (
        request.form.get("guest_count", "").strip()
        or request.form.get("companion", "").strip()
    )

    alcohol_values = [
        value.strip()
        for value in request.form.getlist("alcohol")
        if value.strip()
    ]
    alcohol = ", ".join(alcohol_values)

    transfer = request.form.get("transfer", "").strip()
    children = request.form.get("children", "").strip()
    allergies = request.form.get("allergies", "").strip()

    # Поддерживаем и comment, и старое comments.
    comments = (
        request.form.get("comment", "").strip()
        or request.form.get("comments", "").strip()
    )

    errors = []

    if not name:
        errors.append("Укажите имя и фамилию.")

    if attendance not in {"Приду", "Не смогу прийти"}:
        errors.append("Выберите, сможете ли вы прийти.")

    if companion not in {"Один / одна", "С парой"}:
        errors.append("Укажите, будете вы один/одна или с парой.")

    if errors:
        return jsonify({"ok": False, "message": " ".join(errors)}), 400

    perm_tz = timezone(timedelta(hours=5))
    created_at = datetime.now(perm_tz).strftime("%d.%m.%Y %H:%M")

    conn = None

    try:
        conn = get_db()
        cursor = conn.execute(
            """
            INSERT INTO guests (
                name,
                attendance,
                companion,
                alcohol,
                transfer,
                allergies,
                children,
                comments,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name,
                attendance,
                companion,
                alcohol,
                transfer,
                allergies,
                children,
                comments,
                created_at,
            ),
        )
        conn.commit()
        guest_id = cursor.lastrowid

    except Exception:
        app.logger.exception("Ошибка сохранения RSVP")

        if conn is not None:
            conn.rollback()

        return jsonify(
            {
                "ok": False,
                "message": (
                    "Не удалось сохранить ответ на сервере. "
                    "Посмотрите ошибку в окне, где запущен Flask."
                ),
            }
        ), 500

    finally:
        if conn is not None:
            conn.close()

    return jsonify(
        {
            "ok": True,
            "message": "Спасибо! Ваш ответ сохранён ❤️",
            "guest_id": guest_id,
        }
    )


def admin_key_is_valid():
    expected_key = os.environ.get("ADMIN_KEY", "change-me")
    return request.args.get("key", "") == expected_key


@app.get("/admin")
def admin():
    if not admin_key_is_valid():
        return "Доступ запрещён. Укажите правильный ADMIN_KEY.", 403

    conn = get_db()
    guests = conn.execute(
        "SELECT * FROM guests ORDER BY id DESC"
    ).fetchall()

    stats = conn.execute(
        """
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN attendance = 'Приду' THEN 1 ELSE 0 END) AS coming,
            SUM(CASE WHEN attendance = 'Не смогу прийти' THEN 1 ELSE 0 END) AS not_coming
        FROM guests
        """
    ).fetchone()
    conn.close()

    return render_template(
        "admin.html",
        guests=guests,
        stats=stats,
        admin_key=request.args.get("key", ""),
    )


@app.get("/admin/export.xlsx")
def export_xlsx():
    if not admin_key_is_valid():
        return "Доступ запрещён.", 403

    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM guests ORDER BY id ASC"
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Гости"

    headers = [
        "№",
        "Имя и фамилия",
        "Присутствие",
        "Формат",
        "Алкоголь",
        "Трансфер",
        "Аллергии / питание",
        "Дети",
        "Комментарий",
        "Дата ответа",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="596545")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for index, row in enumerate(rows, start=1):
        ws.append(
            [
                index,
                row["name"],
                row["attendance"],
                row["companion"],
                row["alcohol"],
                row["transfer"],
                row["allergies"],
                row["children"],
                row["comments"],
                row["created_at"],
            ]
        )

    widths = [6, 28, 20, 18, 34, 16, 32, 14, 42, 20]

    for column_number, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column_number)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Гости_Виктор_Валерия.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 5000)),
        debug=True,
    )
